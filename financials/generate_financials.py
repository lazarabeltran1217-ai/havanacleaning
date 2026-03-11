import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLE DEFINITIONS
# ============================================================
DARK_TEAL = "1B4D3E"
MED_TEAL = "2E8B57"
LIGHT_TEAL = "E0F2E9"
GOLD = "D4A017"
LIGHT_GOLD = "FFF8DC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "D9D9D9"

header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color=DARK_TEAL)
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_TEAL)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=MED_TEAL)
category_font = Font(name="Calibri", bold=True, size=11, color=DARK_TEAL)
category_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # yellow = editable
currency_fmt = '$#,##0.00'
pct_fmt = '0.0%'
num_fmt = '#,##0'
thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY)
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_cell(ws, row, col, fmt=None, bold=False, is_input=False):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = Font(name="Calibri", bold=True, size=11)
    if is_input:
        cell.fill = input_fill
    elif row % 2 == 0:
        cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

def style_category_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = category_font
        cell.fill = category_fill
        cell.border = thin_border

def style_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill = PatternFill(start_color=MED_TEAL, end_color=MED_TEAL, fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

def add_title(ws, title, row=1, col=1):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = title_font

def add_subtitle(ws, title, row, col=1):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = subtitle_font

def set_col_widths(ws, widths):
    """widths is dict {col_letter: width} or list of (col_index, width)"""
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

# Helper: cell address string
def cell(col, row):
    """Return cell ref like 'B5'"""
    return f"{get_column_letter(col)}{row}"

def cell_range(col, r1, r2):
    """Return range like 'E4:E50'"""
    c = get_column_letter(col)
    return f"{c}{r1}:{c}{r2}"

# ============================================================
# SHEET 1: INPUTS (Master Control Panel)
# ============================================================
ws_in = wb.active
ws_in.title = "Inputs"
ws_in.sheet_properties.tabColor = DARK_TEAL

add_title(ws_in, "HAVANA CLEANING — EDITABLE INPUTS", 1, 1)
ws_in.merge_cells('A1:C1')
ws_in.cell(row=2, column=1, value="Yellow cells are editable — all other sheets calculate from these values.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

# We'll track every input's row so formulas can reference Inputs!B<row>
# Layout: Column A = Label, Column B = Value, Column C = Notes

r = 4
inp = {}  # dict to store {name: row_number} for formula references

def add_input(label, value, notes="", fmt=None, name=None):
    global r
    ws_in.cell(row=r, column=1, value=label)
    ws_in.cell(row=r, column=2, value=value)
    ws_in.cell(row=r, column=3, value=notes)
    style_cell(ws_in, r, 2, fmt=fmt, is_input=True)
    style_cell(ws_in, r, 1)
    style_cell(ws_in, r, 3)
    if name:
        inp[name] = r
    r += 1

def add_section(title):
    global r
    ws_in.cell(row=r, column=1, value=title)
    style_category_row(ws_in, r, 3)
    r += 1

# --- SERVICE PRICING ---
add_section("SERVICE PRICING")
add_input("Residential Cleaning", 120, "Base price (2-bed/1-bath)", currency_fmt, "price_residential")
add_input("Deep Clean", 220, "Full deep clean", currency_fmt, "price_deep")
add_input("Move-In/Move-Out", 280, "Complete move clean", currency_fmt, "price_move")
add_input("Post-Construction", 350, "Debris & dust removal", currency_fmt, "price_post")
add_input("Airbnb Turnover", 150, "Quick turnaround", currency_fmt, "price_airbnb")
add_input("Green Clean", 140, "Eco-friendly products", currency_fmt, "price_green")
add_input("Recurring Plans", 100, "Discounted recurring", currency_fmt, "price_recurring")
add_input("Commercial (Monthly Contract)", 2500, "Custom quote average", currency_fmt, "price_commercial")
add_input("Avg Add-On Price", 35, "Per add-on service", currency_fmt, "addon_avg")
add_input("Add-On Attach Rate", 0.40, "% of non-commercial jobs", pct_fmt, "addon_rate")
r += 1

# --- LABOR RATES ---
add_section("LABOR RATES & HOURS")
add_input("Entry-Level Cleaner ($/hr)", 20, "Starting rate", currency_fmt, "rate_entry")
add_input("Senior Cleaner ($/hr)", 25, "2+ years experience", currency_fmt, "rate_senior")
add_input("Team Lead ($/hr)", 28, "Supervisory role", currency_fmt, "rate_lead")
add_input("Part-Time Admin ($/hr)", 18, "Scheduling & phones", currency_fmt, "rate_admin")
add_input("Part-Time Bookkeeper ($/hr)", 22, "Accounting support", currency_fmt, "rate_bookkeeper")
add_input("Cleaner Hours/Month", 156, "26 days x 6 productive hrs", num_fmt, "hrs_cleaner")
add_input("Admin Hours/Month", 80, "Part-time ~20 hrs/week", num_fmt, "hrs_admin")
add_input("Bookkeeper Hours/Month", 40, "Part-time ~10 hrs/week", num_fmt, "hrs_bookkeeper")
add_input("Working Days/Month", 26, "Mon-Sat", num_fmt, "work_days")
add_input("Owner Draw Year 1", 0, "Reinvesting Year 1", currency_fmt, "owner_draw_yr1")
add_input("Owner Draw Year 2", 45000, "Annual", currency_fmt, "owner_draw_yr2")
add_input("Owner Draw Year 3", 60000, "Annual", currency_fmt, "owner_draw_yr3")
r += 1

# --- PAYROLL BURDEN ---
add_section("PAYROLL BURDEN")
add_input("Payroll Tax Rate (FICA)", 0.0765, "Social Security + Medicare", pct_fmt, "tax_payroll")
add_input("Workers Comp Rate", 0.045, "FL cleaning industry avg", pct_fmt, "tax_wc")
add_input("Benefits per FT Employee/Mo", 200, "Basic health stipend", currency_fmt, "benefit_per_ft")
r += 1

# --- VEHICLE & OPERATIONS ---
add_section("VEHICLE & OPERATIONS")
add_input("Gas Cost per Day per Vehicle", 25, "Avg 60 miles/day", currency_fmt, "gas_per_day")
add_input("Vehicle Insurance/Month", 300, "Per vehicle", currency_fmt, "vehicle_insurance")
add_input("Vehicle Maintenance/Month", 150, "Per vehicle", currency_fmt, "vehicle_maint")
add_input("Vehicle Parking & Tolls/Month", 50, "Per vehicle", currency_fmt, "vehicle_parking")
add_input("Vehicles Year 1 (end)", 1, "", num_fmt, "vehicles_yr1")
add_input("Vehicles Year 2", 2, "", num_fmt, "vehicles_yr2")
add_input("Vehicles Year 3", 3, "", num_fmt, "vehicles_yr3")
add_input("Supply Cost per Std Job", 8.50, "Consumable supplies", currency_fmt, "supply_per_job")
add_input("Supply Cost per Deep Job", 14, "More product usage", currency_fmt, "supply_per_deep")
add_input("Florida Sales Tax", 0.07, "Collected on services", pct_fmt, "sales_tax")
r += 1

# --- INSURANCE & LICENSING ---
add_section("INSURANCE & LICENSING (Monthly)")
add_input("General Liability Insurance/Mo", 200, "Annual $2,400", currency_fmt, "ins_gl")
add_input("Workers Comp Insurance/Mo", 267, "Annual $3,200", currency_fmt, "ins_wc")
add_input("Surety Bond/Mo", 42, "Annual $500", currency_fmt, "ins_bond")
r += 1

# --- TECHNOLOGY (Monthly) ---
add_section("TECHNOLOGY & SOFTWARE (Monthly)")
add_input("Website Hosting & Domain", 17, "", currency_fmt, "tech_website")
add_input("Scheduling/CRM Software", 50, "", currency_fmt, "tech_crm")
add_input("QuickBooks Online", 45, "", currency_fmt, "tech_qb")
add_input("Business Phone (VoIP)", 35, "", currency_fmt, "tech_phone")
add_input("Internet (home office)", 60, "", currency_fmt, "tech_internet")
add_input("Cell Phone Plans", 85, "", currency_fmt, "tech_cell")
add_input("Stripe Fee Rate", 0.029, "2.9% per transaction", pct_fmt, "stripe_rate")
add_input("Stripe Fee per Transaction", 0.30, "$0.30 per charge", currency_fmt, "stripe_per_txn")
r += 1

# --- OFFICE & ADMIN (Monthly) ---
add_section("OFFICE & ADMIN (Monthly)")
add_input("Office Supplies", 50, "", currency_fmt, "office_supplies")
add_input("Storage Unit Rental", 120, "", currency_fmt, "office_storage")
add_input("Accounting/Bookkeeping Svc", 200, "", currency_fmt, "office_accounting")
add_input("Postage & Shipping", 15, "", currency_fmt, "office_postage")
r += 1

# --- GROWTH RATES ---
add_section("GROWTH ASSUMPTIONS")
add_input("Year 2 Revenue Growth", 1.00, "100% = double Year 1", pct_fmt, "growth_yr2")
add_input("Year 3 Revenue Growth", 0.55, "55% over Year 2", pct_fmt, "growth_yr3")
add_input("Cash Collection Rate (same month)", 0.85, "% collected in service month", pct_fmt, "collect_same")
add_input("Prior Month Collection", 0.10, "% collected next month", pct_fmt, "collect_prior")

set_col_widths(ws_in, {'A': 38, 'B': 18, 'C': 40})

# Store the Inputs sheet reference prefix
I = "Inputs"
def iref(name):
    """Return formula reference to an input cell, e.g. Inputs!$B$5"""
    return f"'{I}'!$B${inp[name]}"

# ============================================================
# SHEET 2: STARTUP COSTS
# ============================================================
ws2 = wb.create_sheet("Startup Costs")
ws2.sheet_properties.tabColor = DARK_TEAL
add_title(ws2, "HAVANA CLEANING — STARTUP COSTS BREAKDOWN", 1, 1)
ws2.merge_cells('A1:E1')
ws2.cell(row=2, column=1, value="Edit QTY and UNIT COST (yellow) — totals auto-calculate.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

headers = ["CATEGORY", "ITEM", "QTY", "UNIT COST", "TOTAL COST"]
r2 = 4
for i, h in enumerate(headers, 1):
    ws2.cell(row=r2, column=i, value=h)
style_header_row(ws2, r2, 5)
r2 += 1
data_start_r2 = r2

startup_items = [
    ("CLEANING EQUIPMENT", None, None, None),
    (None, "Commercial Vacuum Cleaner (Backpack)", 3, 350),
    (None, "Commercial Mop & Bucket System", 3, 85),
    (None, "Microfiber Flat Mop System", 3, 65),
    (None, "Spray Bottles (set of 6)", 4, 18),
    (None, "Telescoping Extension Pole", 3, 35),
    (None, "Scrub Brush Set (multiple sizes)", 4, 22),
    (None, "Squeegee Set (windows)", 3, 28),
    (None, "Dusting Kit (feather duster, microfiber)", 4, 32),
    (None, "Toilet Brush & Caddy Set", 6, 12),
    (None, "Grout Brush Set", 4, 15),
    (None, "Detail Cleaning Brush Kit", 4, 18),
    (None, "Step Ladder (3-step, lightweight)", 2, 75),
    (None, "Caddy/Carrying Tote", 6, 25),
    (None, "Broom & Dustpan Set (commercial grade)", 3, 35),
    (None, "Rubber Gloves (heavy duty, per pair)", 24, 8),
    (None, "Knee Pads", 6, 22),
    (None, "Trash Grabber/Picker Tool", 3, 15),
    ("CLEANING SUPPLIES (3-Month Stock)", None, None, None),
    (None, "All-Purpose Cleaner (gallon concentrate)", 12, 18),
    (None, "Glass Cleaner (gallon concentrate)", 8, 15),
    (None, "Disinfectant Spray (gallon concentrate)", 10, 22),
    (None, "Toilet Bowl Cleaner (bottles)", 24, 5.75),
    (None, "Stainless Steel Cleaner", 8, 9),
    (None, "Wood Floor Cleaner (gallon)", 6, 14),
    (None, "Tile & Grout Cleaner (gallon)", 6, 16),
    (None, "Degreaser (kitchen, gallon)", 6, 18),
    (None, "Microfiber Cloths (packs of 24)", 8, 22),
    (None, "Sponges (packs of 12)", 10, 8),
    (None, "Magic Eraser Pads (packs of 10)", 8, 12),
    (None, "Steel Wool Pads (fine grade, box)", 6, 7),
    (None, "Trash Bags 13-gal (box of 100)", 6, 18),
    (None, "Trash Bags 33-gal (box of 100)", 4, 22),
    (None, "Paper Towels (bulk case of 30 rolls)", 4, 28),
    (None, "Latex/Nitrile Gloves (box of 100)", 12, 11),
    (None, "Shoe Covers (box of 100)", 6, 15),
    (None, "Mop Head Refills (microfiber)", 12, 7.50),
    (None, "Vacuum Bags/Filters (pack)", 6, 15),
    (None, "Air Freshener Spray (bottles)", 12, 6),
    ("GREEN CLEAN PRODUCT LINE", None, None, None),
    (None, "Eco-Friendly All-Purpose Cleaner (gallon)", 6, 24),
    (None, "Plant-Based Disinfectant (gallon)", 4, 28),
    (None, "Non-Toxic Glass Cleaner (gallon)", 4, 20),
    (None, "Biodegradable Trash Bags (box of 100)", 4, 25),
    (None, "Natural Sponges (pack of 6)", 6, 14),
    ("VEHICLES", None, None, None),
    (None, "Used Cargo Van (2020-2022)", 1, 22000),
    (None, "Vehicle Wrap/Branding", 1, 2800),
    (None, "Magnetic Door Signs (backup)", 2, 85),
    (None, "Roof Rack / Interior Shelving", 1, 650),
    (None, "GPS Tracker", 1, 120),
    (None, "Dash Camera", 1, 150),
    (None, "First Aid Kit", 2, 45),
    (None, "Vehicle Insurance Deposit (6 months)", 1, 3600),
    ("TECHNOLOGY & SOFTWARE", None, None, None),
    (None, "Business Laptop (admin)", 1, 800),
    (None, "Tablet for Field Team", 2, 350),
    (None, "Smartphone (business line)", 1, 400),
    (None, "Printer/Scanner", 1, 250),
    (None, "QuickBooks Subscription (Annual)", 1, 540),
    (None, "Business Phone System (VoIP setup)", 1, 200),
    (None, "CRM/Scheduling Software (Annual)", 1, 600),
    ("LEGAL & LICENSING", None, None, None),
    (None, "Florida LLC Formation", 1, 125),
    (None, "Business License (Miami-Dade County)", 1, 50),
    (None, "General Liability Insurance (Annual)", 1, 2400),
    (None, "Workers Compensation Insurance (Annual)", 1, 3200),
    (None, "Surety Bond (Janitorial)", 1, 500),
    (None, "Registered Agent Service (Annual)", 1, 125),
    (None, "Legal Consultation (contracts, waivers)", 1, 1500),
    (None, "Background Check Service (per employee)", 6, 35),
    ("MARKETING & BRANDING", None, None, None),
    (None, "Logo & Brand Design Package", 1, 800),
    (None, "Business Cards (500 count)", 2, 45),
    (None, "Flyers/Door Hangers (5,000 count)", 1, 350),
    (None, "Yard Signs (after-service signs)", 20, 12),
    (None, "Branded Uniforms (polo shirts)", 10, 25),
    (None, "Branded Aprons", 6, 18),
    (None, "Google Ads Initial Budget", 1, 1500),
    (None, "Facebook/Instagram Ads Initial Budget", 1, 1000),
    (None, "Yelp Business Page + Initial Ads", 1, 500),
    (None, "Nextdoor Business Promotion", 1, 200),
    (None, "Referral Program Cards (1,000)", 1, 80),
    (None, "Website Domain & Hosting (Annual)", 1, 200),
    ("OFFICE & ADMIN", None, None, None),
    (None, "Home Office Setup (desk, chair, filing)", 1, 500),
    (None, "Office Supplies (paper, pens, folders)", 1, 150),
    (None, "Storage Unit (3 months deposit)", 3, 120),
    (None, "Safety Data Sheet Binder (OSHA)", 1, 35),
    (None, "Employee Handbook Printing", 10, 8),
    ("CONTINGENCY / EMERGENCY FUND", None, None, None),
    (None, "Emergency Repair Fund", 1, 1500),
    (None, "Working Capital Reserve (2 months)", 1, 12000),
    (None, "Unexpected Expense Buffer", 1, 2000),
]

for cat, item_name, qty, unit_cost in startup_items:
    if cat:  # category header
        ws2.cell(row=r2, column=1, value=cat)
        style_category_row(ws2, r2, 5)
    else:
        ws2.cell(row=r2, column=2, value=item_name)
        ws2.cell(row=r2, column=3, value=qty)
        ws2.cell(row=r2, column=4, value=unit_cost)
        # FORMULA: Total = Qty * Unit Cost
        ws2.cell(row=r2, column=5, value=f"=C{r2}*D{r2}")
        for c in range(1, 6):
            is_inp = c in (3, 4)
            style_cell(ws2, r2, c, fmt=currency_fmt if c in (4,5) else (num_fmt if c == 3 else None), is_input=is_inp)
    r2 += 1

data_end_r2 = r2 - 1
r2 += 1
ws2.cell(row=r2, column=1, value="TOTAL STARTUP INVESTMENT")
ws2.cell(row=r2, column=5, value=f"=SUM(E{data_start_r2}:E{data_end_r2})")
style_total_row(ws2, r2, 5)
ws2.cell(row=r2, column=5).number_format = currency_fmt
startup_total_row = r2

set_col_widths(ws2, {'A': 38, 'B': 48, 'C': 10, 'D': 14, 'E': 16})

# ============================================================
# SHEET 3: MONTHLY SUPPLY COSTS
# ============================================================
ws3 = wb.create_sheet("Monthly Supply Costs")
ws3.sheet_properties.tabColor = MED_TEAL
add_title(ws3, "HAVANA CLEANING — MONTHLY SUPPLY COSTS", 1, 1)
ws3.merge_cells('A1:E1')

headers3 = ["ITEM", "COST/UNIT", "UNITS/MONTH", "MONTHLY COST", "ANNUAL COST"]
r3 = 3
for i, h in enumerate(headers3, 1):
    ws3.cell(row=r3, column=i, value=h)
style_header_row(ws3, r3, 5)
r3 += 1
supply_data_start = r3

supply_items = [
    ("All-Purpose Cleaner (concentrate)", 18.00, 4),
    ("Glass Cleaner (concentrate)", 15.00, 3),
    ("Disinfectant Spray (concentrate)", 22.00, 3),
    ("Toilet Bowl Cleaner", 5.75, 8),
    ("Stainless Steel Cleaner", 9.00, 3),
    ("Wood Floor Cleaner", 14.00, 2),
    ("Tile & Grout Cleaner", 16.00, 2),
    ("Degreaser (kitchen)", 18.00, 2),
    ("Microfiber Cloths (24-pack)", 22.00, 3),
    ("Sponges (12-pack)", 8.00, 3),
    ("Magic Eraser Pads (10-pack)", 12.00, 3),
    ("Trash Bags 13-gal (100ct)", 18.00, 2),
    ("Trash Bags 33-gal (100ct)", 22.00, 1),
    ("Paper Towels (30-roll case)", 28.00, 2),
    ("Latex/Nitrile Gloves (100ct)", 11.00, 4),
    ("Shoe Covers (100ct)", 15.00, 2),
    ("Mop Head Refills", 7.50, 4),
    ("Vacuum Bags/Filters", 15.00, 2),
    ("Air Freshener Spray", 6.00, 4),
    ("Scrub Pads (assorted)", 10.00, 2),
    ("Eco All-Purpose Cleaner", 24.00, 2),
    ("Eco Disinfectant", 28.00, 1),
]

for name, cost, qty in supply_items:
    ws3.cell(row=r3, column=1, value=name)
    ws3.cell(row=r3, column=2, value=cost)
    ws3.cell(row=r3, column=3, value=qty)
    ws3.cell(row=r3, column=4, value=f"=B{r3}*C{r3}")  # Monthly = cost * qty
    ws3.cell(row=r3, column=5, value=f"=D{r3}*12")      # Annual = monthly * 12
    for c in range(1, 6):
        style_cell(ws3, r3, c, fmt=currency_fmt if c in (2,4,5) else (num_fmt if c == 3 else None), is_input=c in (2,3))
    r3 += 1

supply_data_end = r3 - 1
r3 += 1
ws3.cell(row=r3, column=1, value="TOTAL MONTHLY SUPPLY COST")
ws3.cell(row=r3, column=4, value=f"=SUM(D{supply_data_start}:D{supply_data_end})")
ws3.cell(row=r3, column=5, value=f"=SUM(E{supply_data_start}:E{supply_data_end})")
style_total_row(ws3, r3, 5)
ws3.cell(row=r3, column=4).number_format = currency_fmt
ws3.cell(row=r3, column=5).number_format = currency_fmt
supply_total_row = r3

set_col_widths(ws3, {'A': 40, 'B': 14, 'C': 16, 'D': 16, 'E': 16})

# ============================================================
# SHEET 4: EMPLOYEE PROJECTIONS (Year 1 Monthly)
# ============================================================
ws4 = wb.create_sheet("Employee Projections")
ws4.sheet_properties.tabColor = DARK_TEAL
add_title(ws4, "HAVANA CLEANING — YEAR 1 EMPLOYEE & LABOR COST MODEL", 1, 1)
ws4.merge_cells('A1:O1')
ws4.cell(row=2, column=1, value="Edit monthly headcounts (yellow). Wages auto-calculate from Inputs sheet rates.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

# --- HEADCOUNT SECTION ---
add_subtitle(ws4, "MONTHLY HEADCOUNT (editable)", 4, 1)
emp_headers = ["ROLE", "RATE SOURCE"] + [f"M{m+1}" for m in range(12)]
r4 = 6
for i, h in enumerate(emp_headers, 1):
    ws4.cell(row=r4, column=i, value=h)
style_header_row(ws4, r4, 14)
r4 += 1

# Roles with default headcounts and Inputs rate reference
# (role_label, rate_input_name, default_counts_M1-M12)
emp_roles = [
    ("Owner/Operator (cleaning + admin)", "owner_draw_yr1", [1]*12),
    ("Entry-Level Cleaners", "rate_entry", [1, 1, 1, 2, 2, 2, 3, 3, 3, 3, 4, 4]),
    ("Senior Cleaners", "rate_senior", [0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1]),
    ("Part-Time Admin/Scheduler", "rate_admin", [0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1, 1]),
]

role_rows = {}  # {role_index: row}
for idx, (role, rate_name, counts) in enumerate(emp_roles):
    ws4.cell(row=r4, column=1, value=role)
    ws4.cell(row=r4, column=2, value=f"Inputs B{inp[rate_name]}")
    for m in range(12):
        ws4.cell(row=r4, column=3+m, value=counts[m])
        style_cell(ws4, r4, 3+m, fmt=num_fmt, is_input=True)
    style_cell(ws4, r4, 1)
    style_cell(ws4, r4, 2)
    role_rows[idx] = r4
    r4 += 1

# Total headcount row
ws4.cell(row=r4, column=1, value="TOTAL HEADCOUNT")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    first_role_r = role_rows[0]
    last_role_r = role_rows[len(emp_roles)-1]
    ws4.cell(row=r4, column=col, value=f"=SUM({c_letter}{first_role_r}:{c_letter}{last_role_r})")
style_total_row(ws4, r4, 14)
headcount_total_row = r4
r4 += 2

# --- LABOR COST SECTION ---
add_subtitle(ws4, "MONTHLY LABOR COST BREAKDOWN (auto-calculated)", r4, 1)
r4 += 2
cost_headers = ["COST COMPONENT", ""] + [f"M{m+1}" for m in range(12)] + ["ANNUAL"]
for i, h in enumerate(cost_headers, 1):
    ws4.cell(row=r4, column=i, value=h)
style_header_row(ws4, r4, 15)
r4 += 1

# Entry-Level Wages: =headcount * Inputs!rate_entry * Inputs!hrs_cleaner
entry_wages_row = r4
ws4.cell(row=r4, column=1, value="Entry-Level Cleaner Wages")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{role_rows[1]}*{iref('rate_entry')}*{iref('hrs_cleaner')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
# Annual = SUM(M1:M12)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# Senior Wages
senior_wages_row = r4
ws4.cell(row=r4, column=1, value="Senior Cleaner Wages")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{role_rows[2]}*{iref('rate_senior')}*{iref('hrs_cleaner')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# Admin Wages (uses admin hours, not cleaner hours)
admin_wages_row = r4
ws4.cell(row=r4, column=1, value="Part-Time Admin Wages")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{role_rows[3]}*{iref('rate_admin')}*{iref('hrs_admin')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# Gross Wages Subtotal
gross_wages_row = r4
ws4.cell(row=r4, column=1, value="GROSS WAGES SUBTOTAL")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{entry_wages_row}+{c_letter}{senior_wages_row}+{c_letter}{admin_wages_row}")
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_total_row(ws4, r4, 15)
for c in range(3, 16):
    ws4.cell(row=r4, column=c).number_format = currency_fmt
r4 += 1

# Payroll Taxes = Gross Wages * payroll_tax_rate
payroll_tax_row = r4
ws4.cell(row=r4, column=1, value="Payroll Taxes (FICA)")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{gross_wages_row}*{iref('tax_payroll')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# Workers Comp = Gross Wages * wc_rate
wc_row = r4
ws4.cell(row=r4, column=1, value="Workers Compensation")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{gross_wages_row}*{iref('tax_wc')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# Benefits = (Entry count + Senior count) * benefit_per_ft
benefits_row = r4
ws4.cell(row=r4, column=1, value="Employee Benefits")
ws4.cell(row=r4, column=2, value="FT employees only")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    # FT employees = entry cleaners + senior cleaners (not admin who is PT)
    ws4.cell(row=r4, column=col, value=f"=({c_letter}{role_rows[1]}+{c_letter}{role_rows[2]})*{iref('benefit_per_ft')}")
    style_cell(ws4, r4, col, fmt=currency_fmt)
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_cell(ws4, r4, 15, fmt=currency_fmt)
style_cell(ws4, r4, 1)
style_cell(ws4, r4, 2)
r4 += 1

# TOTAL LABOR COST
total_labor_row = r4
ws4.cell(row=r4, column=1, value="TOTAL LABOR COST (All-In)")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws4.cell(row=r4, column=col, value=f"={c_letter}{gross_wages_row}+{c_letter}{payroll_tax_row}+{c_letter}{wc_row}+{c_letter}{benefits_row}")
ws4.cell(row=r4, column=15, value=f"=SUM(C{r4}:N{r4})")
style_total_row(ws4, r4, 15)
for c in range(3, 16):
    ws4.cell(row=r4, column=c).number_format = currency_fmt

set_col_widths(ws4, {get_column_letter(i): (42 if i == 1 else (20 if i == 2 else 14)) for i in range(1, 16)})

# ============================================================
# SHEET 5: REVENUE PROJECTIONS
# ============================================================
ws5 = wb.create_sheet("Revenue Projections")
ws5.sheet_properties.tabColor = GOLD
add_title(ws5, "HAVANA CLEANING — REVENUE PROJECTIONS (YEAR 1)", 1, 1)
ws5.merge_cells('A1:O1')
ws5.cell(row=2, column=1, value="Edit monthly job counts (yellow). Prices pull from Inputs sheet.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

# --- JOB COUNTS ---
add_subtitle(ws5, "MONTHLY JOB COUNTS (editable)", 4, 1)
rev_headers = ["SERVICE TYPE", "PRICE SOURCE"] + [f"M{m+1}" for m in range(12)] + ["ANNUAL JOBS"]
r5 = 6
for i, h in enumerate(rev_headers, 1):
    ws5.cell(row=r5, column=i, value=h)
style_header_row(ws5, r5, 15)
r5 += 1

services_data = [
    ("Residential Cleaning", "price_residential", [8, 10, 14, 18, 22, 25, 28, 30, 33, 36, 40, 45]),
    ("Deep Clean", "price_deep", [2, 3, 4, 5, 6, 7, 8, 8, 9, 10, 11, 12]),
    ("Move-In/Move-Out", "price_move", [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6]),
    ("Post-Construction", "price_post", [0, 1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4]),
    ("Airbnb Turnover", "price_airbnb", [2, 3, 4, 5, 6, 8, 10, 12, 10, 8, 6, 8]),
    ("Green Clean", "price_green", [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6]),
    ("Recurring Plans", "price_recurring", [3, 5, 8, 12, 16, 20, 24, 28, 32, 36, 40, 44]),
    ("Commercial (Monthly)", "price_commercial", [0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3]),
]

svc_rows = {}
for idx, (svc, price_key, jobs) in enumerate(services_data):
    ws5.cell(row=r5, column=1, value=svc)
    ws5.cell(row=r5, column=2, value=f"Inputs B{inp[price_key]}")
    for m in range(12):
        ws5.cell(row=r5, column=3+m, value=jobs[m])
        style_cell(ws5, r5, 3+m, fmt=num_fmt, is_input=True)
    # Annual jobs = SUM
    ws5.cell(row=r5, column=15, value=f"=SUM(C{r5}:N{r5})")
    style_cell(ws5, r5, 15, fmt=num_fmt)
    style_cell(ws5, r5, 1)
    style_cell(ws5, r5, 2)
    svc_rows[idx] = r5
    r5 += 1

# Total Jobs row
total_jobs_row = r5
ws5.cell(row=r5, column=1, value="TOTAL JOBS")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws5.cell(row=r5, column=col, value=f"=SUM({c_letter}{svc_rows[0]}:{c_letter}{svc_rows[len(services_data)-1]})")
ws5.cell(row=r5, column=15, value=f"=SUM(C{r5}:N{r5})")
style_total_row(ws5, r5, 15)
for c in range(3, 16):
    ws5.cell(row=r5, column=c).number_format = num_fmt
r5 += 2

# --- REVENUE BY SERVICE ---
add_subtitle(ws5, "MONTHLY REVENUE BY SERVICE (auto-calculated: jobs x price)", r5, 1)
r5 += 2
rev_headers2 = ["SERVICE TYPE", "UNIT PRICE"] + [f"M{m+1}" for m in range(12)] + ["ANNUAL REV"]
for i, h in enumerate(rev_headers2, 1):
    ws5.cell(row=r5, column=i, value=h)
style_header_row(ws5, r5, 15)
r5 += 1

svc_rev_rows = {}
for idx, (svc, price_key, _) in enumerate(services_data):
    ws5.cell(row=r5, column=1, value=svc)
    # Price formula from Inputs
    ws5.cell(row=r5, column=2, value=f"={iref(price_key)}")
    style_cell(ws5, r5, 2, fmt=currency_fmt)
    for m in range(12):
        col = 3 + m
        c_letter = get_column_letter(col)
        # Revenue = job count (from jobs section above) * price (from Inputs)
        ws5.cell(row=r5, column=col, value=f"={c_letter}{svc_rows[idx]}*{iref(price_key)}")
        style_cell(ws5, r5, col, fmt=currency_fmt)
    ws5.cell(row=r5, column=15, value=f"=SUM(C{r5}:N{r5})")
    style_cell(ws5, r5, 15, fmt=currency_fmt)
    style_cell(ws5, r5, 1)
    svc_rev_rows[idx] = r5
    r5 += 1

# Add-On Revenue row
addon_rev_row = r5
ws5.cell(row=r5, column=1, value="Add-On Services Revenue")
ws5.cell(row=r5, column=2, value=f"={iref('addon_avg')}")
style_cell(ws5, r5, 2, fmt=currency_fmt)
commercial_idx = len(services_data) - 1  # last service is commercial
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    # Add-ons = (total jobs - commercial jobs) * addon_avg * addon_rate
    ws5.cell(row=r5, column=col, value=f"=({c_letter}{total_jobs_row}-{c_letter}{svc_rows[commercial_idx]})*{iref('addon_avg')}*{iref('addon_rate')}")
    style_cell(ws5, r5, col, fmt=currency_fmt)
ws5.cell(row=r5, column=15, value=f"=SUM(C{r5}:N{r5})")
style_cell(ws5, r5, 15, fmt=currency_fmt)
style_cell(ws5, r5, 1)
r5 += 1

# TOTAL GROSS REVENUE
total_rev_row = r5
ws5.cell(row=r5, column=1, value="TOTAL GROSS REVENUE")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws5.cell(row=r5, column=col, value=f"=SUM({c_letter}{svc_rev_rows[0]}:{c_letter}{addon_rev_row})")
ws5.cell(row=r5, column=15, value=f"=SUM(C{r5}:N{r5})")
style_total_row(ws5, r5, 15)
for c in range(3, 16):
    ws5.cell(row=r5, column=c).number_format = currency_fmt

set_col_widths(ws5, {get_column_letter(i): (35 if i == 1 else (16 if i == 2 else 13)) for i in range(1, 16)})

# ============================================================
# SHEET 6: OPERATING EXPENSES
# ============================================================
ws6 = wb.create_sheet("Operating Expenses")
ws6.sheet_properties.tabColor = DARK_TEAL
add_title(ws6, "HAVANA CLEANING — MONTHLY OPERATING EXPENSES (YEAR 1)", 1, 1)
ws6.merge_cells('A1:O1')
ws6.cell(row=2, column=1, value="Edit monthly amounts (yellow). Category totals and grand total auto-calculate.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

opex_headers = ["EXPENSE CATEGORY", "DETAIL"] + [f"M{m+1}" for m in range(12)] + ["ANNUAL"]
r6 = 4
for i, h in enumerate(opex_headers, 1):
    ws6.cell(row=r6, column=i, value=h)
style_header_row(ws6, r6, 15)
r6 += 1

# We'll build opex as categories with items, and track subtotal rows for the grand total
# Items use formulas from Inputs where possible, or are editable values

opex_categories = [
    ("VEHICLE EXPENSES", [
        ("Gas (per vehicle)", None, True),  # formula from Inputs
        ("Vehicle Insurance (per vehicle)", None, True),
        ("Vehicle Maintenance", None, True),
        ("Parking & Tolls", None, True),
    ]),
    ("INSURANCE & LICENSES", [
        ("General Liability Insurance", None, True),
        ("Workers Compensation Insurance", None, True),
        ("Surety Bond", None, True),
    ]),
    ("MARKETING & ADVERTISING", [
        ("Google Ads (PPC)", None, False, [800, 800, 900, 900, 1000, 1000, 1200, 1200, 1000, 1000, 1200, 1500]),
        ("Facebook/Instagram Ads", None, False, [500, 500, 600, 600, 700, 700, 800, 800, 700, 700, 800, 1000]),
        ("Yelp Advertising", None, False, [200]*6 + [300]*6),
        ("Nextdoor / Local Ads", None, False, [100]*6 + [150]*6),
        ("Referral Program Costs", None, False, [0, 25, 50, 75, 100, 125, 150, 175, 200, 225, 250, 275]),
        ("Printed Materials", None, False, [100, 50, 50, 50, 100, 50, 50, 50, 100, 50, 50, 100]),
    ]),
    ("TECHNOLOGY & SOFTWARE", [
        ("Website Hosting", "tech_website", True),
        ("Scheduling/CRM", "tech_crm", True),
        ("QuickBooks", "tech_qb", True),
        ("Business Phone", "tech_phone", True),
        ("Stripe Processing Fees", None, True),  # special formula
        ("Internet", "tech_internet", True),
        ("Cell Phone Plans", "tech_cell", True),
    ]),
    ("CLEANING SUPPLIES", [
        ("Monthly Supply Costs", None, True),  # refs Supply sheet
        ("Equipment Repair/Replace", None, False, [75, 75, 75, 100, 100, 100, 125, 125, 125, 150, 150, 150]),
    ]),
    ("OFFICE & ADMIN", [
        ("Office Supplies", "office_supplies", True),
        ("Storage Unit", "office_storage", True),
        ("Accounting Service", "office_accounting", True),
        ("Postage", "office_postage", True),
    ]),
    ("PROFESSIONAL SERVICES", [
        ("Legal Retainer (quarterly)", None, False, [0, 0, 375, 0, 0, 375, 0, 0, 375, 0, 0, 375]),
        ("CPA/Tax Prep (quarterly)", None, False, [0, 0, 200, 0, 0, 200, 0, 0, 200, 0, 0, 200]),
    ]),
    ("MISCELLANEOUS", [
        ("Background Checks", None, False, [35, 0, 0, 35, 35, 35, 35, 0, 0, 35, 35, 0]),
        ("Training & Development", None, False, [100, 0, 0, 50, 0, 0, 100, 0, 0, 50, 0, 0]),
        ("Uniforms (new hires)", None, False, [50, 0, 0, 50, 25, 25, 50, 0, 0, 50, 25, 0]),
        ("Contingency/Unexpected", None, False, [200]*12),
    ]),
]

subtotal_rows = []
opex_item_rows = []  # track all data rows for the grand total
vehicle_van_count = [1,1,1,1,1,1,1,1,1,1,1,1]  # user can change in Year 1 (1 van all year)

for cat_name, items in opex_categories:
    ws6.cell(row=r6, column=1, value=cat_name)
    style_category_row(ws6, r6, 15)
    cat_start = r6 + 1
    r6 += 1

    for item_data in items:
        detail = item_data[0]
        input_key = item_data[1]
        is_formula = item_data[2]
        values = item_data[3] if len(item_data) > 3 else None

        ws6.cell(row=r6, column=2, value=detail)
        style_cell(ws6, r6, 1)
        style_cell(ws6, r6, 2)

        for m in range(12):
            col = 3 + m
            c_letter = get_column_letter(col)

            if is_formula and input_key:
                # Simple reference to Inputs
                ws6.cell(row=r6, column=col, value=f"={iref(input_key)}")
            elif detail == "Gas (per vehicle)":
                ws6.cell(row=r6, column=col, value=f"={iref('gas_per_day')}*{iref('work_days')}")
            elif detail == "Vehicle Insurance (per vehicle)":
                ws6.cell(row=r6, column=col, value=f"={iref('vehicle_insurance')}")
            elif detail == "Vehicle Maintenance":
                ws6.cell(row=r6, column=col, value=f"={iref('vehicle_maint')}")
            elif detail == "Parking & Tolls":
                ws6.cell(row=r6, column=col, value=f"={iref('vehicle_parking')}")
            elif detail == "General Liability Insurance":
                ws6.cell(row=r6, column=col, value=f"={iref('ins_gl')}")
            elif detail == "Workers Compensation Insurance":
                ws6.cell(row=r6, column=col, value=f"={iref('ins_wc')}")
            elif detail == "Surety Bond":
                ws6.cell(row=r6, column=col, value=f"={iref('ins_bond')}")
            elif detail == "Stripe Processing Fees":
                # Stripe = revenue * stripe_rate + total_jobs * stripe_per_txn
                rev_ref = f"'Revenue Projections'!{c_letter}{total_rev_row}"
                jobs_ref = f"'Revenue Projections'!{c_letter}{total_jobs_row}"
                ws6.cell(row=r6, column=col, value=f"={rev_ref}*{iref('stripe_rate')}+{jobs_ref}*{iref('stripe_per_txn')}")
            elif detail == "Monthly Supply Costs":
                ws6.cell(row=r6, column=col, value=f"='Monthly Supply Costs'!D{supply_total_row}")
            elif values is not None:
                ws6.cell(row=r6, column=col, value=values[m])
                style_cell(ws6, r6, col, fmt=currency_fmt, is_input=True)
                continue  # skip the non-input styling below
            else:
                ws6.cell(row=r6, column=col, value=0)
                style_cell(ws6, r6, col, fmt=currency_fmt, is_input=True)
                continue

            style_cell(ws6, r6, col, fmt=currency_fmt)

        # Annual total
        ws6.cell(row=r6, column=15, value=f"=SUM(C{r6}:N{r6})")
        style_cell(ws6, r6, 15, fmt=currency_fmt)
        opex_item_rows.append(r6)
        r6 += 1

    # Category subtotal
    cat_end = r6 - 1
    ws6.cell(row=r6, column=2, value=f"{cat_name} SUBTOTAL")
    for m in range(12):
        col = 3 + m
        c_letter = get_column_letter(col)
        ws6.cell(row=r6, column=col, value=f"=SUM({c_letter}{cat_start}:{c_letter}{cat_end})")
    ws6.cell(row=r6, column=15, value=f"=SUM(C{r6}:N{r6})")
    style_total_row(ws6, r6, 15)
    for c in range(3, 16):
        ws6.cell(row=r6, column=c).number_format = currency_fmt
    subtotal_rows.append(r6)
    r6 += 1

# GRAND TOTAL
r6 += 1
opex_grand_total_row = r6
ws6.cell(row=r6, column=1, value="TOTAL MONTHLY OPERATING EXPENSES")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    # Sum all subtotal rows
    formula_parts = [f"{c_letter}{sr}" for sr in subtotal_rows]
    ws6.cell(row=r6, column=col, value="=" + "+".join(formula_parts))
ws6.cell(row=r6, column=15, value=f"=SUM(C{r6}:N{r6})")
style_total_row(ws6, r6, 15)
for c in range(3, 16):
    ws6.cell(row=r6, column=c).number_format = currency_fmt

set_col_widths(ws6, {'A': 28, 'B': 38, **{get_column_letter(i): 13 for i in range(3, 16)}})

# ============================================================
# SHEET 7: PROFIT & LOSS STATEMENT (100% formulas)
# ============================================================
ws7 = wb.create_sheet("Profit & Loss")
ws7.sheet_properties.tabColor = GOLD
add_title(ws7, "HAVANA CLEANING — PROFIT & LOSS STATEMENT (YEAR 1)", 1, 1)
ws7.merge_cells('A1:O1')
ws7.cell(row=2, column=1, value="All values auto-calculate from Revenue, Employee, and OpEx sheets.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

pl_headers = ["LINE ITEM", ""] + [f"M{m+1}" for m in range(12)] + ["ANNUAL"]
r7 = 4
for i, h in enumerate(pl_headers, 1):
    ws7.cell(row=r7, column=i, value=h)
style_header_row(ws7, r7, 15)
r7 += 1

# REVENUE
ws7.cell(row=r7, column=1, value="REVENUE")
style_category_row(ws7, r7, 15)
r7 += 1

# Total Revenue = pull from Revenue sheet
pl_rev_row = r7
ws7.cell(row=r7, column=1, value="Total Revenue")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"='Revenue Projections'!{c_letter}{total_rev_row}")
    style_cell(ws7, r7, col, fmt=currency_fmt)
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_cell(ws7, r7, 15, fmt=currency_fmt, bold=True)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 2

# COGS
ws7.cell(row=r7, column=1, value="COST OF GOODS SOLD (Direct Costs)")
style_category_row(ws7, r7, 15)
r7 += 1

# Direct Labor
pl_labor_row = r7
ws7.cell(row=r7, column=1, value="Direct Labor (all-in)")
ws7.cell(row=r7, column=2, value="From Employee Projections")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"='Employee Projections'!{c_letter}{total_labor_row}")
    style_cell(ws7, r7, col, fmt=currency_fmt)
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_cell(ws7, r7, 15, fmt=currency_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 1

# Cleaning Supplies (from supply sheet)
pl_supply_row = r7
ws7.cell(row=r7, column=1, value="Cleaning Supplies")
for m in range(12):
    col = 3 + m
    ws7.cell(row=r7, column=col, value=f"='Monthly Supply Costs'!D{supply_total_row}")
    style_cell(ws7, r7, col, fmt=currency_fmt)
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_cell(ws7, r7, 15, fmt=currency_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 1

# Vehicle Fuel
pl_fuel_row = r7
ws7.cell(row=r7, column=1, value="Vehicle Fuel")
for m in range(12):
    col = 3 + m
    ws7.cell(row=r7, column=col, value=f"={iref('gas_per_day')}*{iref('work_days')}")
    style_cell(ws7, r7, col, fmt=currency_fmt)
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_cell(ws7, r7, 15, fmt=currency_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 1

# TOTAL COGS
pl_cogs_row = r7
ws7.cell(row=r7, column=1, value="TOTAL COGS")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"={c_letter}{pl_labor_row}+{c_letter}{pl_supply_row}+{c_letter}{pl_fuel_row}")
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_total_row(ws7, r7, 15)
for c in range(3, 16):
    ws7.cell(row=r7, column=c).number_format = currency_fmt
r7 += 1

# GROSS PROFIT
pl_gross_row = r7
ws7.cell(row=r7, column=1, value="GROSS PROFIT")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"={c_letter}{pl_rev_row}-{c_letter}{pl_cogs_row}")
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_total_row(ws7, r7, 15)
for c in range(3, 16):
    ws7.cell(row=r7, column=c).number_format = currency_fmt
r7 += 1

# Gross Margin %
pl_gm_row = r7
ws7.cell(row=r7, column=1, value="Gross Margin %")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"=IF({c_letter}{pl_rev_row}=0,0,{c_letter}{pl_gross_row}/{c_letter}{pl_rev_row})")
    style_cell(ws7, r7, col, fmt=pct_fmt)
ws7.cell(row=r7, column=15, value=f"=IF(O{pl_rev_row}=0,0,O{pl_gross_row}/O{pl_rev_row})")
style_cell(ws7, r7, 15, fmt=pct_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 2

# OPERATING EXPENSES (SG&A)
ws7.cell(row=r7, column=1, value="OPERATING EXPENSES (SG&A)")
style_category_row(ws7, r7, 15)
r7 += 1

pl_opex_row = r7
ws7.cell(row=r7, column=1, value="Total Operating Expenses")
ws7.cell(row=r7, column=2, value="From OpEx sheet")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"='Operating Expenses'!{c_letter}{opex_grand_total_row}")
    style_cell(ws7, r7, col, fmt=currency_fmt)
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_cell(ws7, r7, 15, fmt=currency_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 2

# NET OPERATING INCOME
pl_net_row = r7
ws7.cell(row=r7, column=1, value="NET OPERATING INCOME")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"={c_letter}{pl_gross_row}-{c_letter}{pl_opex_row}")
ws7.cell(row=r7, column=15, value=f"=SUM(C{r7}:N{r7})")
style_total_row(ws7, r7, 15)
for c in range(3, 16):
    ws7.cell(row=r7, column=c).number_format = currency_fmt
r7 += 1

# Net Margin %
pl_nm_row = r7
ws7.cell(row=r7, column=1, value="Net Margin %")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws7.cell(row=r7, column=col, value=f"=IF({c_letter}{pl_rev_row}=0,0,{c_letter}{pl_net_row}/{c_letter}{pl_rev_row})")
    style_cell(ws7, r7, col, fmt=pct_fmt)
ws7.cell(row=r7, column=15, value=f"=IF(O{pl_rev_row}=0,0,O{pl_net_row}/O{pl_rev_row})")
style_cell(ws7, r7, 15, fmt=pct_fmt)
style_cell(ws7, r7, 1)
style_cell(ws7, r7, 2)
r7 += 2

# CUMULATIVE NET INCOME
pl_cum_row = r7
ws7.cell(row=r7, column=1, value="CUMULATIVE NET INCOME")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    if m == 0:
        ws7.cell(row=r7, column=col, value=f"={c_letter}{pl_net_row}")
    else:
        prev_letter = get_column_letter(col - 1)
        ws7.cell(row=r7, column=col, value=f"={prev_letter}{pl_cum_row}+{c_letter}{pl_net_row}")
style_total_row(ws7, r7, 15)
for c in range(3, 15):
    ws7.cell(row=r7, column=c).number_format = currency_fmt
# Annual = last month cumulative
ws7.cell(row=r7, column=15, value=f"=N{pl_cum_row}")
ws7.cell(row=r7, column=15).number_format = currency_fmt

set_col_widths(ws7, {'A': 35, 'B': 28, **{get_column_letter(i): 13 for i in range(3, 16)}})

# ============================================================
# SHEET 8: CASH FLOW
# ============================================================
ws8 = wb.create_sheet("Cash Flow")
ws8.sheet_properties.tabColor = MED_TEAL
add_title(ws8, "HAVANA CLEANING — CASH FLOW PROJECTIONS (YEAR 1)", 1, 1)
ws8.merge_cells('A1:O1')

cf_headers = ["CASH FLOW ITEM", ""] + [f"M{m+1}" for m in range(12)] + ["ANNUAL"]
r8 = 3
for i, h in enumerate(cf_headers, 1):
    ws8.cell(row=r8, column=i, value=h)
style_header_row(ws8, r8, 15)
r8 += 1

# BEGINNING CASH BALANCE
cf_begin_row = r8
ws8.cell(row=r8, column=1, value="BEGINNING CASH BALANCE")
# M1 = 0 (before investment), M2+ = prior month ending
ws8.cell(row=r8, column=3, value=0)
for m in range(1, 12):
    col = 3 + m
    prev_letter = get_column_letter(col - 1)
    # Will reference ending balance row (defined below)
    # Placeholder — we'll fix after defining ending row
style_total_row(ws8, r8, 15)
for c in range(3, 16):
    ws8.cell(row=r8, column=c).number_format = currency_fmt
r8 += 1

# CASH INFLOWS
ws8.cell(row=r8, column=1, value="CASH INFLOWS")
style_category_row(ws8, r8, 15)
r8 += 1

# Cash collected from services
cf_collected_row = r8
ws8.cell(row=r8, column=1, value="")
ws8.cell(row=r8, column=2, value="Cash from Services")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    prev_letter = get_column_letter(col - 1) if m > 0 else None
    # Same-month collection + prior month collection
    same_month = f"'Profit & Loss'!{c_letter}{pl_rev_row}*{iref('collect_same')}"
    if m == 0:
        ws8.cell(row=r8, column=col, value=f"={same_month}")
    else:
        prior_month = f"'Profit & Loss'!{prev_letter}{pl_rev_row}*{iref('collect_prior')}"
        ws8.cell(row=r8, column=col, value=f"={same_month}+{prior_month}")
    style_cell(ws8, r8, col, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Owner Capital Investment (M1 = startup total, rest = 0, editable)
cf_invest_row = r8
ws8.cell(row=r8, column=2, value="Owner Capital Investment")
ws8.cell(row=r8, column=3, value=f"='Startup Costs'!E{startup_total_row}+5000")  # startup costs + small buffer
for m in range(1, 12):
    ws8.cell(row=r8, column=3+m, value=0)
for m in range(12):
    style_cell(ws8, r8, 3+m, fmt=currency_fmt, is_input=(m > 0))
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Total Cash Inflows
cf_inflows_row = r8
ws8.cell(row=r8, column=1, value="TOTAL CASH INFLOWS")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"={c_letter}{cf_collected_row}+{c_letter}{cf_invest_row}")
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_total_row(ws8, r8, 15)
for c in range(3, 16):
    ws8.cell(row=r8, column=c).number_format = currency_fmt
r8 += 1

# CASH OUTFLOWS
ws8.cell(row=r8, column=1, value="CASH OUTFLOWS")
style_category_row(ws8, r8, 15)
r8 += 1

# Labor costs
cf_labor_out = r8
ws8.cell(row=r8, column=2, value="Labor Costs (all-in)")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"='Profit & Loss'!{c_letter}{pl_labor_row}")
    style_cell(ws8, r8, col, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Operating Expenses
cf_opex_out = r8
ws8.cell(row=r8, column=2, value="Operating Expenses")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"='Operating Expenses'!{c_letter}{opex_grand_total_row}")
    style_cell(ws8, r8, col, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Supply Costs
cf_supply_out = r8
ws8.cell(row=r8, column=2, value="Cleaning Supplies")
for m in range(12):
    col = 3 + m
    ws8.cell(row=r8, column=col, value=f"='Monthly Supply Costs'!D{supply_total_row}")
    style_cell(ws8, r8, col, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Startup Equipment (M1 only)
cf_startup_out = r8
ws8.cell(row=r8, column=2, value="Startup Costs (Month 1)")
ws8.cell(row=r8, column=3, value=f"='Startup Costs'!E{startup_total_row}")
for m in range(1, 12):
    ws8.cell(row=r8, column=3+m, value=0)
for m in range(12):
    style_cell(ws8, r8, 3+m, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Sales Tax (quarterly)
cf_tax_out = r8
ws8.cell(row=r8, column=2, value="Sales Tax Remittance (quarterly)")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    if m in [2, 5, 8, 11]:  # quarterly months
        # Sum revenue for the quarter * sales tax rate
        q_start_col = get_column_letter(col - 2)
        q_mid_col = get_column_letter(col - 1)
        ws8.cell(row=r8, column=col, value=f"=('Profit & Loss'!{q_start_col}{pl_rev_row}+'Profit & Loss'!{q_mid_col}{pl_rev_row}+'Profit & Loss'!{c_letter}{pl_rev_row})*{iref('sales_tax')}")
    else:
        ws8.cell(row=r8, column=col, value=0)
    style_cell(ws8, r8, col, fmt=currency_fmt)
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_cell(ws8, r8, 15, fmt=currency_fmt)
style_cell(ws8, r8, 1)
style_cell(ws8, r8, 2)
r8 += 1

# Total Cash Outflows
cf_outflows_row = r8
ws8.cell(row=r8, column=1, value="TOTAL CASH OUTFLOWS")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"={c_letter}{cf_labor_out}+{c_letter}{cf_opex_out}+{c_letter}{cf_supply_out}+{c_letter}{cf_startup_out}+{c_letter}{cf_tax_out}")
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_total_row(ws8, r8, 15)
for c in range(3, 16):
    ws8.cell(row=r8, column=c).number_format = currency_fmt
r8 += 1

# NET CASH FLOW
cf_netcf_row = r8
ws8.cell(row=r8, column=1, value="NET CASH FLOW")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"={c_letter}{cf_inflows_row}-{c_letter}{cf_outflows_row}")
ws8.cell(row=r8, column=15, value=f"=SUM(C{r8}:N{r8})")
style_total_row(ws8, r8, 15)
for c in range(3, 16):
    ws8.cell(row=r8, column=c).number_format = currency_fmt
r8 += 1

# ENDING CASH BALANCE
cf_ending_row = r8
ws8.cell(row=r8, column=1, value="ENDING CASH BALANCE")
for m in range(12):
    col = 3 + m
    c_letter = get_column_letter(col)
    ws8.cell(row=r8, column=col, value=f"={c_letter}{cf_begin_row}+{c_letter}{cf_netcf_row}")
ws8.cell(row=r8, column=15, value=f"=N{cf_ending_row}")  # ending = last month
style_total_row(ws8, r8, 15)
for c in range(3, 16):
    ws8.cell(row=r8, column=c).number_format = currency_fmt

# NOW fix beginning balance row to reference ending balance of prior month
for m in range(1, 12):
    col = 3 + m
    prev_letter = get_column_letter(col - 1)
    ws8.cell(row=cf_begin_row, column=col, value=f"={prev_letter}{cf_ending_row}")

set_col_widths(ws8, {'A': 32, 'B': 35, **{get_column_letter(i): 13 for i in range(3, 16)}})

# ============================================================
# SHEET 9: BREAK-EVEN ANALYSIS
# ============================================================
ws9 = wb.create_sheet("Break-Even Analysis")
ws9.sheet_properties.tabColor = GOLD
add_title(ws9, "HAVANA CLEANING — BREAK-EVEN ANALYSIS", 1, 1)
ws9.merge_cells('A1:D1')
ws9.cell(row=2, column=1, value="All calculations reference Inputs and Operating Expenses sheets.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

r9 = 4
be_headers = ["COMPONENT", "DESCRIPTION", "MONTHLY AMOUNT", "NOTES"]
for i, h in enumerate(be_headers, 1):
    ws9.cell(row=r9, column=i, value=h)
style_header_row(ws9, r9, 4)
r9 += 1

# FIXED COSTS
ws9.cell(row=r9, column=1, value="MONTHLY FIXED COSTS")
style_category_row(ws9, r9, 4)
r9 += 1
fixed_start = r9

fixed_items = [
    ("Base Labor (owner + 1 cleaner)", "Min staffing", f"={iref('rate_entry')}*{iref('hrs_cleaner')}", "1 cleaner at entry rate"),
    ("Payroll Taxes on Base Labor", "FICA", f"=C{r9}*{iref('tax_payroll')}", "Applied to cleaner wages"),  # will be r9+1
    ("Workers Comp on Base Labor", "WC", None, "Applied to cleaner wages"),
    ("Benefits (1 FT employee)", "Health stipend", f"={iref('benefit_per_ft')}", "Per FT employee"),
    ("Vehicle Insurance", "1 van", f"={iref('vehicle_insurance')}", "Monthly premium"),
    ("Vehicle Maintenance", "1 van", f"={iref('vehicle_maint')}", "Monthly avg"),
    ("General Liability Insurance", "Monthly", f"={iref('ins_gl')}", "Required"),
    ("Workers Comp Insurance", "Monthly", f"={iref('ins_wc')}", "Required"),
    ("Surety Bond", "Monthly", f"={iref('ins_bond')}", "Janitorial bond"),
    ("Technology & Software", "Monthly total", f"={iref('tech_website')}+{iref('tech_crm')}+{iref('tech_qb')}+{iref('tech_phone')}+{iref('tech_internet')}+{iref('tech_cell')}", "All subscriptions"),
    ("Office & Admin", "Monthly total", f"={iref('office_supplies')}+{iref('office_storage')}+{iref('office_accounting')}+{iref('office_postage')}", "Overhead"),
    ("Marketing (minimum)", "Base ad spend", 1200, "Editable minimum"),
    ("Contingency", "Buffer", 200, "Editable buffer"),
]

for idx, (comp, desc, amount, notes) in enumerate(fixed_items):
    ws9.cell(row=r9, column=1, value=comp)
    ws9.cell(row=r9, column=2, value=desc)
    if comp == "Payroll Taxes on Base Labor":
        ws9.cell(row=r9, column=3, value=f"=C{r9-1}*{iref('tax_payroll')}")
    elif comp == "Workers Comp on Base Labor":
        ws9.cell(row=r9, column=3, value=f"=C{r9-2}*{iref('tax_wc')}")
    elif amount is not None:
        ws9.cell(row=r9, column=3, value=amount)
    ws9.cell(row=r9, column=4, value=notes)
    is_inp = isinstance(amount, (int, float)) and comp in ("Marketing (minimum)", "Contingency")
    for c in range(1, 5):
        style_cell(ws9, r9, c, fmt=currency_fmt if c == 3 else None, is_input=(is_inp and c == 3))
    r9 += 1

fixed_end = r9 - 1
# Total Fixed
be_fixed_total_row = r9
ws9.cell(row=r9, column=1, value="TOTAL FIXED COSTS / MONTH")
ws9.cell(row=r9, column=3, value=f"=SUM(C{fixed_start}:C{fixed_end})")
style_total_row(ws9, r9, 4)
ws9.cell(row=r9, column=3).number_format = currency_fmt
r9 += 2

# VARIABLE COSTS PER JOB
ws9.cell(row=r9, column=1, value="VARIABLE COSTS PER JOB")
style_category_row(ws9, r9, 4)
r9 += 1
var_start = r9

var_items = [
    ("Cleaning Supplies", "Consumables", f"={iref('supply_per_job')}", "Per standard job"),
    ("Gas/Travel", "Per job portion", f"={iref('gas_per_day')}/3", "$25/day ÷ 3 jobs/day"),
    ("Stripe Processing", "2.9% + $0.30", f"=150*{iref('stripe_rate')}+{iref('stripe_per_txn')}", "On ~$150 avg job"),
    ("Equipment Wear & Tear", "Amortized", 2.00, "Replacement amortization"),
]

for comp, desc, amount, notes in var_items:
    ws9.cell(row=r9, column=1, value=comp)
    ws9.cell(row=r9, column=2, value=desc)
    ws9.cell(row=r9, column=3, value=amount)
    ws9.cell(row=r9, column=4, value=notes)
    is_inp = isinstance(amount, (int, float))
    for c in range(1, 5):
        style_cell(ws9, r9, c, fmt=currency_fmt if c == 3 else None, is_input=(is_inp and c == 3))
    r9 += 1

var_end = r9 - 1
be_var_total_row = r9
ws9.cell(row=r9, column=1, value="TOTAL VARIABLE COST / JOB")
ws9.cell(row=r9, column=3, value=f"=SUM(C{var_start}:C{var_end})")
style_total_row(ws9, r9, 4)
ws9.cell(row=r9, column=3).number_format = currency_fmt
r9 += 2

# BREAK-EVEN CALCULATION
ws9.cell(row=r9, column=1, value="BREAK-EVEN CALCULATION")
style_category_row(ws9, r9, 4)
r9 += 1

# Average Revenue per Job = Year 1 total revenue / total jobs
be_avg_rev_row = r9
ws9.cell(row=r9, column=1, value="Avg Revenue per Job")
ws9.cell(row=r9, column=3, value=f"='Profit & Loss'!O{pl_rev_row}/'Revenue Projections'!O{total_jobs_row}")
ws9.cell(row=r9, column=4, value="Year 1 weighted average")
for c in range(1, 5):
    style_cell(ws9, r9, c, fmt=currency_fmt if c == 3 else None)
r9 += 1

be_contrib_row = r9
ws9.cell(row=r9, column=1, value="Contribution Margin / Job")
ws9.cell(row=r9, column=3, value=f"=C{be_avg_rev_row}-C{be_var_total_row}")
ws9.cell(row=r9, column=4, value="Revenue - Variable Cost")
for c in range(1, 5):
    style_cell(ws9, r9, c, fmt=currency_fmt if c == 3 else None)
r9 += 1

be_jobs_row = r9
ws9.cell(row=r9, column=1, value="BREAK-EVEN JOBS / MONTH")
ws9.cell(row=r9, column=3, value=f"=C{be_fixed_total_row}/C{be_contrib_row}")
ws9.cell(row=r9, column=4, value="Fixed Costs ÷ Contribution Margin")
style_total_row(ws9, r9, 4)
ws9.cell(row=r9, column=3).number_format = '0.0'
r9 += 1

be_rev_row = r9
ws9.cell(row=r9, column=1, value="BREAK-EVEN REVENUE / MONTH")
ws9.cell(row=r9, column=3, value=f"=C{be_jobs_row}*C{be_avg_rev_row}")
style_total_row(ws9, r9, 4)
ws9.cell(row=r9, column=3).number_format = currency_fmt
r9 += 1

be_daily_row = r9
ws9.cell(row=r9, column=1, value="BREAK-EVEN JOBS / DAY")
ws9.cell(row=r9, column=3, value=f"=C{be_jobs_row}/{iref('work_days')}")
ws9.cell(row=r9, column=4, value="Jobs ÷ working days/month")
style_total_row(ws9, r9, 4)
ws9.cell(row=r9, column=3).number_format = '0.0'

set_col_widths(ws9, {'A': 34, 'B': 22, 'C': 18, 'D': 40})

# ============================================================
# SHEET 10: 3-YEAR SUMMARY
# ============================================================
ws10 = wb.create_sheet("3-Year Summary")
ws10.sheet_properties.tabColor = DARK_TEAL
add_title(ws10, "HAVANA CLEANING — 3-YEAR FINANCIAL SUMMARY", 1, 1)
ws10.merge_cells('A1:E1')
ws10.cell(row=2, column=1, value="Year 1 from P&L. Years 2-3 auto-scale by growth rates on Inputs sheet.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

r10 = 4
h10 = ["LINE ITEM", "YEAR 1", "YEAR 2", "YEAR 3", "3-YEAR TOTAL"]
for i, h in enumerate(h10, 1):
    ws10.cell(row=r10, column=i, value=h)
style_header_row(ws10, r10, 5)
r10 += 1

# Year 1 Revenue = P&L annual
yr_rev_row = r10
ws10.cell(row=r10, column=1, value="Total Revenue")
ws10.cell(row=r10, column=2, value=f"='Profit & Loss'!O{pl_rev_row}")  # Year 1
ws10.cell(row=r10, column=3, value=f"=B{r10}*(1+{iref('growth_yr2')})")  # Year 2
ws10.cell(row=r10, column=4, value=f"=C{r10}*(1+{iref('growth_yr3')})")  # Year 3
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=currency_fmt if c >= 2 else None)
r10 += 1

# Revenue Growth %
ws10.cell(row=r10, column=1, value="Revenue Growth %")
ws10.cell(row=r10, column=2, value="—")
ws10.cell(row=r10, column=3, value=f"={iref('growth_yr2')}")
ws10.cell(row=r10, column=4, value=f"={iref('growth_yr3')}")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=pct_fmt if c in (3, 4) else None)
r10 += 1

# Year 1 COGS
yr_cogs_row = r10
ws10.cell(row=r10, column=1, value="Total COGS")
ws10.cell(row=r10, column=2, value=f"='Profit & Loss'!O{pl_cogs_row}")
# Year 2/3 COGS scales with revenue but improves margin slightly (economies of scale)
ws10.cell(row=r10, column=3, value=f"=B{yr_cogs_row}/B{yr_rev_row}*C{yr_rev_row}*0.95")  # 5% efficiency gain
ws10.cell(row=r10, column=4, value=f"=C{yr_cogs_row}/C{yr_rev_row}*D{yr_rev_row}*0.93")  # 7% efficiency gain
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=currency_fmt if c >= 2 else None)
r10 += 1

# Gross Profit
yr_gross_row = r10
ws10.cell(row=r10, column=1, value="GROSS PROFIT")
for c in range(2, 5):
    c_letter = get_column_letter(c)
    ws10.cell(row=r10, column=c, value=f"={c_letter}{yr_rev_row}-{c_letter}{yr_cogs_row}")
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
style_total_row(ws10, r10, 5)
for c in range(2, 6):
    ws10.cell(row=r10, column=c).number_format = currency_fmt
r10 += 1

# Gross Margin %
ws10.cell(row=r10, column=1, value="Gross Margin %")
for c in range(2, 5):
    c_letter = get_column_letter(c)
    ws10.cell(row=r10, column=c, value=f"=IF({c_letter}{yr_rev_row}=0,0,{c_letter}{yr_gross_row}/{c_letter}{yr_rev_row})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=pct_fmt if c >= 2 else None)
r10 += 1

# Operating Expenses
yr_opex_row = r10
ws10.cell(row=r10, column=1, value="Operating Expenses (SG&A)")
ws10.cell(row=r10, column=2, value=f"='Profit & Loss'!O{pl_opex_row}")
ws10.cell(row=r10, column=3, value=f"=B{r10}*1.3")  # 30% increase Year 2
ws10.cell(row=r10, column=4, value=f"=C{r10}*1.2")  # 20% increase Year 3
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=currency_fmt if c >= 2 else None)
r10 += 1

# Owner Draw
yr_draw_row = r10
ws10.cell(row=r10, column=1, value="Owner Draw / Salary")
ws10.cell(row=r10, column=2, value=f"={iref('owner_draw_yr1')}")
ws10.cell(row=r10, column=3, value=f"={iref('owner_draw_yr2')}")
ws10.cell(row=r10, column=4, value=f"={iref('owner_draw_yr3')}")
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=currency_fmt if c >= 2 else None)
r10 += 1

# Net Operating Income
yr_net_row = r10
ws10.cell(row=r10, column=1, value="NET OPERATING INCOME")
for c in range(2, 5):
    c_letter = get_column_letter(c)
    ws10.cell(row=r10, column=c, value=f"={c_letter}{yr_gross_row}-{c_letter}{yr_opex_row}-{c_letter}{yr_draw_row}")
ws10.cell(row=r10, column=5, value=f"=SUM(B{r10}:D{r10})")
style_total_row(ws10, r10, 5)
for c in range(2, 6):
    ws10.cell(row=r10, column=c).number_format = currency_fmt
r10 += 1

# Net Margin %
ws10.cell(row=r10, column=1, value="Net Margin %")
for c in range(2, 5):
    c_letter = get_column_letter(c)
    ws10.cell(row=r10, column=c, value=f"=IF({c_letter}{yr_rev_row}=0,0,{c_letter}{yr_net_row}/{c_letter}{yr_rev_row})")
for c in range(1, 6):
    style_cell(ws10, r10, c, fmt=pct_fmt if c >= 2 else None)
r10 += 2

# Startup Investment
ws10.cell(row=r10, column=1, value="STARTUP INVESTMENT REQUIRED")
ws10.cell(row=r10, column=2, value=f"='Startup Costs'!E{startup_total_row}")
style_total_row(ws10, r10, 5)
ws10.cell(row=r10, column=2).number_format = currency_fmt

set_col_widths(ws10, {'A': 32, 'B': 18, 'C': 18, 'D': 18, 'E': 18})

# ============================================================
# SHEET 11: EQUIPMENT CHECKLIST
# ============================================================
ws11 = wb.create_sheet("Equipment Checklist")
ws11.sheet_properties.tabColor = MED_TEAL
add_title(ws11, "HAVANA CLEANING — EQUIPMENT & SUPPLY CHECKLIST", 1, 1)
ws11.merge_cells('A1:F1')

r11 = 3
h11 = ["ITEM", "CATEGORY", "QTY", "UNIT COST", "TOTAL", "REPLACE FREQUENCY"]
for i, h in enumerate(h11, 1):
    ws11.cell(row=r11, column=i, value=h)
style_header_row(ws11, r11, 6)
r11 += 1
equip_start = r11

checklist = [
    ("Commercial Backpack Vacuum", "Core Equipment", 3, 350, "Every 3-5 years"),
    ("HEPA Filter Vacuum", "Core Equipment", 1, 450, "Every 3-5 years"),
    ("Commercial Mop & Bucket System", "Core Equipment", 3, 85, "Every 2-3 years"),
    ("Microfiber Flat Mop System", "Core Equipment", 3, 65, "Every 1-2 years"),
    ("Steam Cleaner (handheld)", "Core Equipment", 2, 120, "Every 3-4 years"),
    ("Extension Pole (telescoping)", "Core Equipment", 3, 35, "Every 3-5 years"),
    ("Step Ladder (aluminum)", "Core Equipment", 2, 75, "Every 5+ years"),
    ("Wet/Dry Shop Vac", "Core Equipment", 1, 180, "Every 3-5 years"),
    ("Spray Bottles (color-coded)", "Cleaning Tools", 24, 3, "Every 6-12 months"),
    ("Scrub Brush Set", "Cleaning Tools", 4, 22, "Every 6-12 months"),
    ("Squeegee (professional)", "Cleaning Tools", 3, 28, "Every 1-2 years"),
    ("Grout Brush Set", "Cleaning Tools", 4, 15, "Every 6-12 months"),
    ("Detail Brush Kit", "Cleaning Tools", 4, 18, "Every 3-6 months"),
    ("Toilet Brush & Caddy", "Cleaning Tools", 6, 12, "Every 3-6 months"),
    ("Dusting Wand (extendable)", "Cleaning Tools", 4, 15, "Every 6-12 months"),
    ("Broom & Dustpan (commercial)", "Cleaning Tools", 3, 35, "Every 1-2 years"),
    ("Putty Knife/Scraper Set", "Cleaning Tools", 3, 12, "Every 1-2 years"),
    ("Razor Blade Scraper", "Cleaning Tools", 3, 8, "Replace blades monthly"),
    ("Cleaning Caddy/Tote", "Organization", 6, 25, "Every 2-3 years"),
    ("Rolling Utility Cart", "Organization", 2, 85, "Every 3-5 years"),
    ("Rubber Gloves (pairs)", "Safety/PPE", 24, 8, "Monthly"),
    ("Nitrile Gloves (100ct box)", "Safety/PPE", 12, 11, "Monthly"),
    ("Knee Pads (gel)", "Safety/PPE", 6, 22, "Every 6-12 months"),
    ("Safety Glasses", "Safety/PPE", 6, 8, "Every 6-12 months"),
    ("Shoe Covers (100ct box)", "Safety/PPE", 6, 15, "Monthly"),
    ("N95 Masks", "Safety/PPE", 12, 3, "Weekly"),
    ("First Aid Kit", "Safety/PPE", 2, 45, "Restock quarterly"),
    ("Wet Floor Signs", "Safety/PPE", 3, 12, "Every 2+ years"),
    ("Microfiber Cloths (24-pack)", "Textiles", 8, 22, "Every 3-6 months"),
    ("Glass Cloths (lint-free)", "Textiles", 6, 8, "Every 3-6 months"),
    ("Scrub Sponges (12-pack)", "Textiles", 10, 8, "Monthly"),
    ("Magic Eraser Pads (10-pack)", "Textiles", 8, 12, "Monthly"),
    ("Mop Head Refills", "Textiles", 12, 7.50, "Every 2-4 weeks"),
    ("Vacuum Bags/Filters", "Textiles", 6, 15, "Monthly"),
    ("Cargo Van Interior Shelving", "Vehicle", 1, 650, "One-time"),
    ("Magnetic Door Signs", "Vehicle", 2, 85, "Every 2-3 years"),
    ("Vehicle Wrap (full)", "Vehicle", 1, 2800, "Every 3-5 years"),
    ("GPS Tracker", "Vehicle", 1, 120, "Every 3-5 years"),
    ("Dash Camera", "Vehicle", 1, 150, "Every 2-3 years"),
]

for item, cat, qty, cost, freq in checklist:
    ws11.cell(row=r11, column=1, value=item)
    ws11.cell(row=r11, column=2, value=cat)
    ws11.cell(row=r11, column=3, value=qty)
    ws11.cell(row=r11, column=4, value=cost)
    ws11.cell(row=r11, column=5, value=f"=C{r11}*D{r11}")  # FORMULA
    ws11.cell(row=r11, column=6, value=freq)
    for c in range(1, 7):
        style_cell(ws11, r11, c, fmt=currency_fmt if c in (4, 5) else (num_fmt if c == 3 else None), is_input=c in (3, 4))
    r11 += 1

equip_end = r11 - 1
r11 += 1
ws11.cell(row=r11, column=1, value="TOTAL EQUIPMENT INVESTMENT")
ws11.cell(row=r11, column=5, value=f"=SUM(E{equip_start}:E{equip_end})")
style_total_row(ws11, r11, 6)
ws11.cell(row=r11, column=5).number_format = currency_fmt

set_col_widths(ws11, {'A': 38, 'B': 18, 'C': 10, 'D': 14, 'E': 14, 'F': 24})

# ============================================================
# SHEET 12: KPI DASHBOARD
# ============================================================
ws12 = wb.create_sheet("KPI Dashboard")
ws12.sheet_properties.tabColor = GOLD
add_title(ws12, "HAVANA CLEANING — KEY PERFORMANCE INDICATORS", 1, 1)
ws12.merge_cells('A1:C1')
ws12.cell(row=2, column=1, value="All KPIs auto-calculate from other sheets.").font = Font(name="Calibri", italic=True, size=10, color=MED_TEAL)

r12 = 4
h12 = ["KPI METRIC", "YEAR 1 VALUE", "FORMULA SOURCE"]
for i, h in enumerate(h12, 1):
    ws12.cell(row=r12, column=i, value=h)
style_header_row(ws12, r12, 3)
r12 += 1

# Financial KPIs
ws12.cell(row=r12, column=1, value="FINANCIAL KPIs")
style_category_row(ws12, r12, 3)
r12 += 1

kpi_items = [
    ("Total Annual Revenue", f"='Profit & Loss'!O{pl_rev_row}", currency_fmt, "P&L Annual Revenue"),
    ("Total COGS", f"='Profit & Loss'!O{pl_cogs_row}", currency_fmt, "P&L Annual COGS"),
    ("Gross Profit", f"='Profit & Loss'!O{pl_gross_row}", currency_fmt, "Revenue - COGS"),
    ("Gross Margin %", f"='Profit & Loss'!O{pl_gm_row}", pct_fmt, "Gross Profit / Revenue"),
    ("Total Operating Expenses", f"='Profit & Loss'!O{pl_opex_row}", currency_fmt, "P&L Annual SG&A"),
    ("Net Operating Income", f"='Profit & Loss'!O{pl_net_row}", currency_fmt, "Gross Profit - OpEx"),
    ("Net Margin %", f"='Profit & Loss'!O{pl_nm_row}", pct_fmt, "Net Income / Revenue"),
    ("Ending Cash Balance", f"='Cash Flow'!N{cf_ending_row}", currency_fmt, "Cash Flow M12 ending"),
]

for label, formula, fmt, source in kpi_items:
    ws12.cell(row=r12, column=1, value=label)
    ws12.cell(row=r12, column=2, value=formula)
    ws12.cell(row=r12, column=3, value=source)
    for c in range(1, 4):
        style_cell(ws12, r12, c, fmt=fmt if c == 2 else None)
    r12 += 1

r12 += 1
ws12.cell(row=r12, column=1, value="OPERATIONAL KPIs")
style_category_row(ws12, r12, 3)
r12 += 1

op_kpis = [
    ("Total Jobs Completed (Year 1)", f"='Revenue Projections'!O{total_jobs_row}", num_fmt, "Revenue sheet annual jobs"),
    ("Average Revenue per Job", f"='Profit & Loss'!O{pl_rev_row}/'Revenue Projections'!O{total_jobs_row}", currency_fmt, "Revenue / Total Jobs"),
    ("Jobs per Working Day (avg)", f"='Revenue Projections'!O{total_jobs_row}/({iref('work_days')}*12)", '0.0', "Annual jobs / working days"),
    ("Break-Even Jobs/Month", f"='Break-Even Analysis'!C{be_jobs_row}", '0.0', "From break-even sheet"),
    ("Break-Even Revenue/Month", f"='Break-Even Analysis'!C{be_rev_row}", currency_fmt, "From break-even sheet"),
    ("Year-End Headcount", f"='Employee Projections'!N{headcount_total_row}", num_fmt, "M12 headcount"),
    ("Revenue per Employee", f"='Profit & Loss'!O{pl_rev_row}/'Employee Projections'!N{headcount_total_row}", currency_fmt, "Revenue / headcount"),
    ("Startup Investment", f"='Startup Costs'!E{startup_total_row}", currency_fmt, "Startup Costs sheet total"),
]

for label, formula, fmt, source in op_kpis:
    ws12.cell(row=r12, column=1, value=label)
    ws12.cell(row=r12, column=2, value=formula)
    ws12.cell(row=r12, column=3, value=source)
    for c in range(1, 4):
        style_cell(ws12, r12, c, fmt=fmt if c == 2 else None)
    r12 += 1

r12 += 1
ws12.cell(row=r12, column=1, value="3-YEAR OUTLOOK")
style_category_row(ws12, r12, 3)
r12 += 1

outlook_kpis = [
    ("Year 2 Projected Revenue", f"='3-Year Summary'!C{yr_rev_row}", currency_fmt, "Year 1 x (1 + growth rate)"),
    ("Year 3 Projected Revenue", f"='3-Year Summary'!D{yr_rev_row}", currency_fmt, "Year 2 x (1 + growth rate)"),
    ("Year 2 Net Income", f"='3-Year Summary'!C{yr_net_row}", currency_fmt, "3-Year Summary"),
    ("Year 3 Net Income", f"='3-Year Summary'!D{yr_net_row}", currency_fmt, "3-Year Summary"),
    ("3-Year Total Revenue", f"='3-Year Summary'!E{yr_rev_row}", currency_fmt, "Sum of 3 years"),
    ("3-Year Total Net Income", f"='3-Year Summary'!E{yr_net_row}", currency_fmt, "Sum of 3 years"),
]

for label, formula, fmt, source in outlook_kpis:
    ws12.cell(row=r12, column=1, value=label)
    ws12.cell(row=r12, column=2, value=formula)
    ws12.cell(row=r12, column=3, value=source)
    for c in range(1, 4):
        style_cell(ws12, r12, c, fmt=fmt if c == 2 else None)
    r12 += 1

set_col_widths(ws12, {'A': 36, 'B': 22, 'C': 36})

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\beltr\Havana Cleaning\financials\Havana_Cleaning_Financial_Projections.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"\nSheets: {wb.sheetnames}")
print(f"\nAll values are FORMULA-DRIVEN:")
print(f"  - Inputs sheet: {len(inp)} editable parameters (yellow cells)")
print(f"  - Startup Costs: QTY x COST formulas + SUM total")
print(f"  - Supply Costs: COST x QTY formulas + SUM totals")
print(f"  - Employee Projections: headcount x rate x hours formulas")
print(f"  - Revenue Projections: job counts x prices from Inputs")
print(f"  - Operating Expenses: references Inputs + SUM subtotals")
print(f"  - P&L: 100% cross-sheet formulas")
print(f"  - Cash Flow: references P&L + collection rate formulas")
print(f"  - Break-Even: fixed/variable cost formulas")
print(f"  - 3-Year Summary: growth rate formulas from Inputs")
print(f"  - Equipment Checklist: QTY x COST formulas")
print(f"  - KPI Dashboard: cross-sheet formula references")
